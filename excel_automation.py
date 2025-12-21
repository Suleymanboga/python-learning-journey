"""Excel Automation project with chart feature"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']
    price_multiplier = float(input("Please enter the multiplier for your prices: "))
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        updated_price = cell.value * price_multiplier
        updated_price_cell = sheet.cell(row, 4)
        updated_price_cell.value = updated_price
    values = Reference(sheet,
            min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "f2")

    workbook.save("updated_" + filename)
if __name__ == "__main__":
    filename_ = input("Please enter the name of your excel file with .xlsx in the end: ")
    process_workbook(filename_)
